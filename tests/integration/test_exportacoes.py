"""
O número de sócio precisa sair nas quatro saídas de dados: CSV, Excel,
webhook e busca do admin. Este arquivo cobre as três que passam por HTTP.
"""

import csv
import io


async def _criar_associado(
    db_session, nome, cpf, numero, telefone, modalidades=None, outros=None
):
    from app.models import (
        Associado,
        AssociadoDepartamento,
        Candidato,
        Departamento,
        Preferencia,
        Resposta,
    )

    candidato = Candidato(nome="Fulano", apelido="Fu", ativo=True)
    db_session.add(candidato)
    await db_session.commit()
    await db_session.refresh(candidato)

    associado = Associado(
        nome=nome,
        cpf=cpf,
        numero_socio=numero,
        telefone=telefone,
        departamento_outros=outros,
        aceite_lgpd=True,
    )
    db_session.add(associado)
    await db_session.commit()
    await db_session.refresh(associado)

    db_session.add(Resposta(associado_id=associado.id, candidato_id=candidato.id))
    db_session.add(
        Preferencia(associado_id=associado.id, candidato_preferido_id=candidato.id)
    )

    for ordem, nome_mod in enumerate(modalidades or [], start=1):
        dep = Departamento(nome=nome_mod, ordem=ordem, ativo=True)
        db_session.add(dep)
        await db_session.commit()
        await db_session.refresh(dep)
        db_session.add(
            AssociadoDepartamento(associado_id=associado.id, departamento_id=dep.id)
        )

    await db_session.commit()
    return associado


async def _preparar_voto_export(
    client, db_session, valid_cpf, numero_socio, read_otp_code, telefone
):
    """Igual ao de test_departamentos.py — repetido aqui de propósito para
    que cada arquivo de teste seja legível sozinho."""
    from app.models import Candidato

    candidato = Candidato(nome="Fulano", apelido="Fu", ativo=True)
    db_session.add(candidato)
    await db_session.commit()
    await db_session.refresh(candidato)

    res = await client.post(
        "/api/survey/register",
        json={
            "nome": "Socio Export",
            "cpf": valid_cpf(),
            "telefone": telefone,
            "numero_socio": numero_socio(),
            "titular": True,
            "recaptcha_token": "",
        },
    )
    token = res.json()["session_token"]
    codigo = read_otp_code(telefone)
    await client.post(
        "/api/survey/verify-otp",
        json={"session_token": token, "telefone": telefone, "codigo": codigo},
    )
    return token, candidato.id


class TestExportacaoCSV:
    async def test_csv_tem_coluna_de_numero_de_socio(
        self, client, db_session, admin_token, valid_cpf
    ):
        await _criar_associado(db_session, "Socio CSV", valid_cpf(), "0042", "11944440001")

        res = await client.get(
            "/api/admin/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.status_code == 200

        linhas = list(csv.reader(io.StringIO(res.text)))
        assert "Nº Sócio" in linhas[0]

        indice = linhas[0].index("Nº Sócio")
        assert linhas[1][indice] == "0042"


class TestExportacaoExcel:
    async def test_excel_tem_coluna_de_numero_de_socio(
        self, client, db_session, admin_token, valid_cpf
    ):
        from openpyxl import load_workbook

        await _criar_associado(db_session, "Socio XLS", valid_cpf(), "0043", "11944440002")

        res = await client.get(
            "/api/admin/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.status_code == 200

        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        cabecalho = [c.value for c in ws[1]]
        assert "Nº Sócio" in cabecalho

        indice = cabecalho.index("Nº Sócio")
        assert ws[2][indice].value == "0043"


class TestBuscaDoAdmin:
    async def test_busca_devolve_numero_de_socio(
        self, client, db_session, admin_token, valid_cpf
    ):
        cpf = valid_cpf()
        await _criar_associado(db_session, "Socio Busca", cpf, "0044", "11944440003")

        res = await client.get(
            f"/api/admin/search?cpf={cpf}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.status_code == 200
        assert res.json()[0]["numero_socio"] == "0044"


class TestModalidadesNasExportacoes:
    async def test_csv_tem_as_duas_colunas(
        self, client, db_session, admin_token, valid_cpf
    ):
        await _criar_associado(
            db_session,
            "Socio Mod",
            valid_cpf(),
            "0050",
            "11922220001",
            modalidades=["Natação", "Sauna"],
            outros=None,
        )

        res = await client.get(
            "/api/admin/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        linhas = list(csv.reader(io.StringIO(res.text)))
        assert "Modalidades" in linhas[0]
        assert "Outros (descrição)" in linhas[0]
        assert linhas[1][linhas[0].index("Modalidades")] == "Natação, Sauna"

    async def test_excel_tem_as_duas_colunas(
        self, client, db_session, admin_token, valid_cpf
    ):
        from openpyxl import load_workbook

        await _criar_associado(
            db_session,
            "Socio Mod XLS",
            valid_cpf(),
            "0051",
            "11922220002",
            modalidades=["Judô"],
            outros="Xadrez",
        )

        res = await client.get(
            "/api/admin/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        ws = load_workbook(io.BytesIO(res.content)).active
        cabecalho = [c.value for c in ws[1]]
        assert ws[2][cabecalho.index("Modalidades")].value == "Judô"
        assert ws[2][cabecalho.index("Outros (descrição)")].value == "Xadrez"

    async def test_busca_do_admin_devolve_as_modalidades(
        self, client, db_session, admin_token, valid_cpf
    ):
        cpf = valid_cpf()
        await _criar_associado(
            db_session,
            "Socio Busca Mod",
            cpf,
            "0052",
            "11922220003",
            modalidades=["Piscina"],
            outros=None,
        )

        res = await client.get(
            f"/api/admin/search?cpf={cpf}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.json()[0]["departamentos"] == ["Piscina"]
        assert res.json()[0]["departamento_outros"] == ""

    async def test_csv_respeita_a_ordem_e_nao_o_nome(
        self, client, db_session, admin_token, valid_cpf
    ):
        """Mesmo motivo do teste equivalente em test_departamentos.py: sob a
        collation do banco, ordenar por nome devolveria outra sequência.
        "Zumba" (ordem 1) antes de "Atletismo" (ordem 2) torna a divergência
        óbvia — uma regressão para sort por nome devolveria a ordem trocada."""
        await _criar_associado(
            db_session,
            "Socio Mod Ordem CSV",
            valid_cpf(),
            "0053",
            "11922220006",
            modalidades=["Zumba", "Atletismo"],
            outros=None,
        )

        res = await client.get(
            "/api/admin/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        linhas = list(csv.reader(io.StringIO(res.text)))
        assert linhas[1][linhas[0].index("Modalidades")] == "Zumba, Atletismo"

    async def test_excel_respeita_a_ordem_e_nao_o_nome(
        self, client, db_session, admin_token, valid_cpf
    ):
        """Mesmo motivo do teste equivalente em test_departamentos.py."""
        from openpyxl import load_workbook

        await _criar_associado(
            db_session,
            "Socio Mod Ordem XLS",
            valid_cpf(),
            "0054",
            "11922220007",
            modalidades=["Zumba", "Atletismo"],
            outros=None,
        )

        res = await client.get(
            "/api/admin/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        ws = load_workbook(io.BytesIO(res.content)).active
        cabecalho = [c.value for c in ws[1]]
        assert ws[2][cabecalho.index("Modalidades")].value == "Zumba, Atletismo"

    async def test_busca_do_admin_respeita_a_ordem_e_nao_o_nome(
        self, client, db_session, admin_token, valid_cpf
    ):
        """Mesmo motivo do teste equivalente em test_departamentos.py."""
        cpf = valid_cpf()
        await _criar_associado(
            db_session,
            "Socio Mod Ordem Busca",
            cpf,
            "0055",
            "11922220008",
            modalidades=["Zumba", "Atletismo"],
            outros=None,
        )

        res = await client.get(
            f"/api/admin/search?cpf={cpf}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.json()[0]["departamentos"] == ["Zumba", "Atletismo"]


class TestPayloadDoWebhook:
    """O payload nunca teve teste automatizado — o campo numero_socio só foi
    conferido à mão, com um receptor HTTP. Como esta mudança mexe nele,
    entra a cobertura."""

    async def test_payload_serializado_tem_as_modalidades(
        self, client, db_session, valid_cpf, numero_socio, read_otp_code, departamentos
    ):
        import json

        from sqlalchemy import select

        from app.models import WebhookLog

        deps = await departamentos()
        token, cid = await _preparar_voto_export(
            client, db_session, valid_cpf, numero_socio, read_otp_code, "11922220004"
        )

        await client.post(
            "/api/survey/submit",
            json={
                "session_token": token,
                "candidatos_ids": [cid],
                "candidato_preferido_id": cid,
                "departamentos_ids": [deps[0].id, deps[-1].id],
                "departamento_outros": "Xadrez",
                "aceite_lgpd": True,
            },
        )

        log = (await db_session.execute(select(WebhookLog))).scalars().first()
        payload = json.loads(log.payload)
        assert payload["departamentos"] == [deps[0].nome, deps[-1].nome]
        assert payload["departamento_outros"] == "Xadrez"

    async def test_sem_outros_o_campo_vai_como_string_vazia(
        self, client, db_session, valid_cpf, numero_socio, read_otp_code, departamentos
    ):
        """Nunca null: o n8n trata campo ausente e campo nulo de formas
        diferentes."""
        import json

        from sqlalchemy import select

        from app.models import WebhookLog

        deps = await departamentos()
        token, cid = await _preparar_voto_export(
            client, db_session, valid_cpf, numero_socio, read_otp_code, "11922220005"
        )

        await client.post(
            "/api/survey/submit",
            json={
                "session_token": token,
                "candidatos_ids": [cid],
                "candidato_preferido_id": cid,
                "departamentos_ids": [deps[0].id],
                "aceite_lgpd": True,
            },
        )

        log = (await db_session.execute(select(WebhookLog))).scalars().first()
        assert json.loads(log.payload)["departamento_outros"] == ""

    async def test_payload_respeita_a_ordem_e_nao_o_nome(
        self, client, db_session, valid_cpf, numero_socio, read_otp_code
    ):
        """Mesmo motivo do teste equivalente em test_departamentos.py: sob a
        collation do banco, ordenar por nome devolveria outra sequência. Não
        usa a fixture `departamentos` porque os nomes dela ("Modalidade 1",
        "Modalidade 2"...) coincidem em ordem alfabética e em `ordem` — não
        provariam nada. `departamentos_ids` é enviado na ordem alfabética
        (Atletismo antes de Zumba), que também é o oposto de `ordem`: só o
        resultado correto (ordenado por `ordem`) bate com o esperado."""
        import json

        from sqlalchemy import select

        from app.models import Departamento, WebhookLog

        db_session.add_all(
            [
                Departamento(nome="Zumba", ordem=1, ativo=True),
                Departamento(nome="Atletismo", ordem=2, ativo=True),
            ]
        )
        await db_session.commit()
        deps = (
            (
                await db_session.execute(
                    select(Departamento).order_by(Departamento.ordem)
                )
            )
            .scalars()
            .all()
        )
        zumba, atletismo = deps[0], deps[1]

        token, cid = await _preparar_voto_export(
            client, db_session, valid_cpf, numero_socio, read_otp_code, "11922220010"
        )

        await client.post(
            "/api/survey/submit",
            json={
                "session_token": token,
                "candidatos_ids": [cid],
                "candidato_preferido_id": cid,
                "departamentos_ids": [atletismo.id, zumba.id],
                "aceite_lgpd": True,
            },
        )

        log = (await db_session.execute(select(WebhookLog))).scalars().first()
        payload = json.loads(log.payload)
        assert payload["departamentos"] == ["Zumba", "Atletismo"]

    async def test_deserialize_payload_tolera_campos_historicos(self):
        """O worker de retry (SurveyService.retry_pending_webhook) re-hidrata
        payloads GRAVADOS em webhook_logs chamando deserialize_payload, sem
        try/except em volta. Linhas gravadas antes de numero_socio,
        departamentos e departamento_outros existirem não carregam esses
        campos — e sem default, WebhookPayload(**json.loads(...)) levanta
        ValidationError ANTES de log.tentativas ser incrementado, então uma
        única linha antiga trava o worker de retry para sempre (ele
        reprocessa a mesma linha em loop e nunca chega às pendências
        legítimas atrás dela). Todo campo adicionado a WebhookPayload depois
        da v1 precisa de default, para sempre, por causa disso."""
        import json

        from app.integrations.webhook import WebhookService

        formato_original = {
            "nome": "Maria Santos",
            "cpf": "12345678909",
            "telefone": "11999998888",
            "candidatos": ["Fulano"],
            "preferido": "Fulano",
            "aceite_lgpd": True,
            "data": "2024-01-01T00:00:00",
        }
        payload = WebhookService.deserialize_payload(json.dumps(formato_original))
        assert payload.numero_socio == ""
        assert payload.departamentos == []
        assert payload.departamento_outros == ""
        assert payload.titular is False

        formato_pos_numero_socio = {**formato_original, "numero_socio": "1234"}
        payload = WebhookService.deserialize_payload(
            json.dumps(formato_pos_numero_socio)
        )
        assert payload.numero_socio == "1234"
        assert payload.departamentos == []
        assert payload.departamento_outros == ""
        assert payload.titular is False

        # Terceira forma histórica: já tem modalidades, ainda não tem titular.
        # É a forma que está sendo gravada em produção neste exato momento,
        # ou seja, a que o worker vai reencontrar depois do deploy desta
        # feature. Foi por não escrever este caso na branch anterior que o
        # mesmo bug apareceu duas vezes.
        formato_pos_departamentos = {
            **formato_pos_numero_socio,
            "departamentos": ["Natação"],
            "departamento_outros": "",
        }
        payload = WebhookService.deserialize_payload(
            json.dumps(formato_pos_departamentos)
        )
        assert payload.departamentos == ["Natação"]
        assert payload.titular is False
